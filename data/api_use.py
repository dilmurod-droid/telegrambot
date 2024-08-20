import json
import os
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

# from handlers.users.test_comp import JWT_SECRET

# from data.create_token import verify_or_create_token
# from openpyxl.styles import Font, Alignment
BASE_URL='https://dilmurod0887.pythonanywhere.com/api/'
# i_e = {
#     'id': 'fdg',
#     'end_date': 'dgfh',
# }
# def create_profile(user_id,phone_number,name,username):
#     url=f"{BASE_URL}/bot-users/create"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     post=requests.post(url,data={'user_id':user_id,'name':name,'phone_number':phone_number,'username':username},headers=headers)
#     data=json.loads(post.text)
#     return data
# def data_time():
#     current_date_str = datetime.now().strftime("%Y-%m-%d")
#
#     parsed_date = datetime.strptime(current_date_str, "%Y-%m-%d")
#
#     formatted_date = parsed_date.strftime("%Y-%m-%d")
#     return f"{formatted_date}"
# def user_created(user_id):
#     url=f"{BASE_URL}/bot-users/users"
#     token = verify_or_create_token('dli','dli','adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response=requests.get(url,headers=headers)
#     data=json.loads(response.text)
#     for user in data:
#         if user['user_id']==user_id:
#             return True
#
# def get_tariffs():
#     url = f"{BASE_URL}/tariffs/"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.get(url, headers=headers)
#     data = json.loads(response.text)
#     return data
# def create_payments(user,amount,success,check_image_id):
#     url = f"{BASE_URL}/payments/create"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.post(url=url,data={'user':user,'amount':amount,'success':success,'check_image_id':check_image_id},headers=headers)
#     data = json.loads(response.text)
#     return data
# def get_users(username):
#     url = f"{BASE_URL}/bot-users/users"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.get(url,headers=headers)
#     data = json.loads(response.text)
#     for data in data:
#         if data['username']==username:
#             return data
# def get_all_users():
#     url = f"{BASE_URL}/bot-users/users"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.get(url, headers=headers)
#     data = json.loads(response.text)
#     return data
# def get_tariffs_by_name(name):
#     url = f"{BASE_URL}/tariffs/"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.get(url, headers=headers)
#     data = json.loads(response.text)
#     for data in data:
#         if data['name']==name:
#             return data
# def get_tariffs_by_id(id):
#     url = f"{BASE_URL}/tariffs/"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.get(url, headers=headers)
#     data = json.loads(response.text)
#     for data in data:
#         if data['id']==id:
#             return data['name']
# def get_payment_name(id:int):
#     url = f"{BASE_URL}/payments/"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.get(url,headers=headers)
#     data = json.loads(response.text)
#     for data in data:
#         if data['user']==id:
#             return data
# def orders_create(user,tariff,payment,days,is_active):
#     url = f"{BASE_URL}/order/create"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.post(url=url,data={'user':user,'tariff':tariff,'payment':payment,'days':days,'is_active':is_active},headers=headers)
#     data = json.loads(response.text)
#     return data
#
# def get_ordered_tariffs(user):
#     url = f"{BASE_URL}/orders/"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.get(url, headers=headers)
#     data = json.loads(response.text)
#     for data in data:
#         if data['user'] == user:
#             if data['end_date']==data_time():
#                 url = f"{BASE_URL}/orders/{data['id']}/"
#                 data = {
#                     "is_active": False
#                 }
#                 response = requests.patch(url, json=data)
#                 return response.text
# def get_ordered_tariff(user):
#     global i_e
#     url = f"{BASE_URL}/orders/"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     response = requests.get(url, headers=headers)
#     data = json.loads(response.text)
#     for data in data:
#         if data['user']==user:
#             return {'id':data['tariff'],'end_date':data['end_date']}
#
# # .xlsx
# data=[
#     {
#
# }
# ]
# id=0
# end_date="0000-00-00"
# def t_name():
#     global data,id,end_date
#     users=get_all_users()
#     for i in users:
#         result=get_ordered_tariff(i['id'])
#         if result:
#             id = result['id']
#             end_date=result['end_date']
#         else:
#             id=0
#             end_date="0000-00-00"
#         tariff = get_tariffs_by_id(id)
#         data.append({'name': i['name'], 'phone_number': i['phone_number'],'ordered_tariff':tariff if tariff else 'Tarif sotib olmagan','end_date':end_date})
#     return data
#
#
# def create_or_update_file(file_path,data):
#     """
#     Fayl mavjud bo'lsa, ichidagi ma'lumotlarni o'chiradi va yangilaydi yoki yangi fayl yaratadi.
#     """
#     if os.path.exists(file_path):
#         # Fayl mavjud bo'lsa, uni ochamiz
#         wb = load_workbook(file_path)
#         ws = wb.active
#
#         # Mavjud ma'lumotlarni o'qish
#         existing_data = set()
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             existing_data.add(tuple(row))  # Har bir qatorni tuple sifatida saqlaymiz
#     else:
#         # Agar fayl mavjud bo'lmasa, yangi yaratamiz
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Yangilangan Ma'lumotlar"
#         # Ustun nomlarini yozish
#         headers = ["Ismi", "Telefon raqami", "Sotib olgan tarifi","Tariff tugash vaqti"]
#         ws.append(headers)
#         existing_data = set()
#
#     for entry in data:
#         row = (entry.get("name", ""), entry.get("phone_number", ""), entry.get("ordered_tariff", ""),entry.get("end_date",""))
#         if row not in existing_data:
#             ws.append(row)
#             existing_data.add(row)
#
#     # Ustun kengliklarini sozlash
#     for col in ws.columns:
#         max_length = 0
#         column = col[0].column_letter  # Ustun harfi
#         for cell in col:
#             try:
#                 if cell.value is not None and len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = (max_length + 2)
#         ws.column_dimensions[column].width = adjusted_width
#
#     # Excel faylni saqlash
#     wb.save(file_path)
#
# def create_file_user():
#     t_name()
#     create_or_update_file('users.xlsx',data)
#
# def cruser(name,description):
#     url=f"https://dilmurod0887.pythonanywhere.com/api/users/"
#     token = verify_or_create_token('dli', 'dli', 'adsfsg')['access']
#     headers = {
#         'Authorization': f'Bearer {token}'
#     }
#     post=requests.post(url,data={'name':name,'description':description},headers=headers)
#     data=json.loads(post.text)
#     return data
#
#
# print(cruser('20+60','bu zor'))
def login(username,password):
    url=f"{BASE_URL}/login/"
    post=requests.post(url,data={'username':username,'password':password})
    data=json.loads(post.text)
    return data

import json




def read_from_json(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)
def write_to_json(file_path, new_data):
    try:
        # Try to read existing data from file
        with open(file_path, 'r') as file:
            existing_data = json.load(file)
    except FileNotFoundError:
        # If file does not exist, initialize with an empty list
        existing_data = []
    except json.JSONDecodeError:
        # If file is empty or corrupted, initialize with an empty list
        existing_data = []
    except ValueError as e:
        # If existing data is not a list, handle the error
        print(f"Error: {e}")
        existing_data = []

    # Combine existing data with new data
    if isinstance(existing_data, list):
        existing_data.append(new_data)
    else:
        existing_data = new_data

    # Write the updated data back to file
    with open(file_path, 'w') as file:
        json.dump(existing_data, file, indent=4)

def create_resualt(user_id,test_id,token):
    url=f"{BASE_URL}/bot-users/create"
    headers = {
        'Authorization': f'Bearer {token}'
    }
    post=requests.post(url,data={'user':user_id,'test':test_id},headers=headers)
    data=json.loads(post.text)
    return data
def get_tests(token):
    url = f"{BASE_URL}/tests/"
    headers = {
        'Authorization': f'Bearer {token}'
    }
    response = requests.get(url, headers=headers)
    return response.json()
def get_questions(token):
    url = f"{BASE_URL}/tests/"
    headers = {
        'Authorization': f'Bearer {token}'
    }
    response = requests.get(url, headers=headers)
    return response.json()
def get_results(token):
    url = f"{BASE_URL}/results/"
    headers = {
        'Authorization': f'Bearer {token}'
    }
    response = requests.get(url, headers=headers)
    return response.json()
def get_profile(token):
    url = f"{BASE_URL}/profile/"
    headers = {
        'Authorization': f'Bearer {token}'
    }
    response = requests.get(url, headers=headers)
    return response.json()


import requests


def get_test_and_user_details():
    # token=read_from_json()['others']['access']
    headers = {'Authorization': f'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzI1MDM0NDEyLCJpYXQiOjE3MjQxNzA0MTIsImp0aSI6ImU4NDlkZTk0MzkzNDQ4ZTQ5YmJhMjMxNDM0NTM2ZDExIiwidXNlcl9pZCI6MTZ9.SfaRcF_Z53t8mnndPCnVN7nrS0RCy6ONiXNv9YdUg04'}
    results_response = requests.get(f"{BASE_URL}/users/resualts", headers=headers)

    if results_response.status_code == 200:
        results = results_response.json()
        formatted_results = []
        for result in results:
            # Extract test and user IDs from each result
            test_id = result.get("test")
            user_id = result.get("user")

            # Endpoint to get test and user details
            test_url = f"{BASE_URL}/tests/{test_id}"
            user_url = f"{BASE_URL}/users/{user_id}"

            # Get test name
            test_response = requests.get(test_url, headers=headers)
            if test_response.status_code == 200:
                test_data = test_response.json()
                test_name = test_data.get("name", "Unknown Test")
            else:
                test_name = "Unknown Test"

            # Get user username
            user_response = requests.get(user_url, headers=headers)
            if user_response.status_code == 200:
                user_data = user_response.json()
                username = user_data.get("username", "Unknown User")
            else:
                username = "Unknown User"

            # Print results
            formatted_results.append({
                "Test_Name": test_name,
                "Username": username,
                "Score": result.get("score"),
                "Incorrect_Answers": result.get("incorrect_answers"),
                "Correct_Answers": result.get("correct_answers"),
                "Failed": result.get("failed")
            })
        return formatted_results
    else:
        print(f"Failed to retrieve the results.{results_response.text}")


# Example usage
# print(read_from_json('users.json'))
# print(write_to_json('users.json',{'id': 6667155546, 'others': {'user_id': 16, 'refresh': 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ0b2tlbl90eXBlIjoicmVmcmVzaCIsImV4cCI6MTczMjgyMjMwNSwiaWF0IjoxNzI0MTgyMzA1LCJqdGkiOiI0ZDVlNmMyMDYwNmM0YTMxYmMyOTEwNjlkM2I4M2U0MCIsInVzZXJfaWQiOjE2fQ.tG4tL3MAMWpOdxGdbVmVCwez9VYfv_iJ1VzIC-vMoxo', 'access': 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzI1MDQ2MzA1LCJpYXQiOjE3MjQxODIzMDUsImp0aSI6ImFjZThmYWRkMGI1NTQ0NmQ4ZDgzYmRmMmQ4NGU2ZmVkIiwidXNlcl9pZCI6MTZ9.ftDkXS5kZo7AP0wzSJvNOIIf43li0Jsu4QzJmTrWWWE', 'role': 'mentor'}}))
# data = get_profile('eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzI1MDM0NDEyLCJpYXQiOjE3MjQxNzA0MTIsImp0aSI6ImU4NDlkZTk0MzkzNDQ4ZTQ5YmJhMjMxNDM0NTM2ZDExIiwidXNlcl9pZCI6MTZ9.SfaRcF_Z53t8mnndPCnVN7nrS0RCy6ONiXNv9YdUg04')
# print(data)
# data={'id':123,'others':login('admin','admin')}

# print(write_to_json('users.json',data))
# print(read_from_json('users.json')['id'])